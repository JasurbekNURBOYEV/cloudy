import json
import os

import pyowm
from django.conf import settings
from openpyxl import Workbook

from core import models
from core.utils.time import timestamp_to_datetime, get_now


def load_cities_from_json():
    path = settings.BASE_DIR / 'files/city.list.json'
    limit = settings.CITY_LIMIT
    city_added_count = 0
    country_added_count = 0
    with open(path, 'r') as f:
        data = json.loads(f.read())
        total_count = len(data)
        print(f'Total {total_count} cities detected in json...')
        c = 1
        existing_cities = models.City.all()
        if existing_cities.count() >= limit:
            print(f'Can not load, city limit reached, you can only load {limit} cities')
            to_be_deleted = existing_cities[limit:]
            print(f'{to_be_deleted.count()} cities are being deleted...')
            for city in to_be_deleted:
                city.delete()
            return
        total_count = limit - existing_cities.count()

        for i in range(total_count):
            city = data[i]
            print(f'Processing {c}/{total_count} ({c/total_count*100:.4f}%): {city["name"]}')
            country_instance = models.Country.get(code=city['country'])
            if not country_instance:
                country_added_count += 1
                country_instance, _ = models.Country.objects.get_or_create(code=city['country'])
            city_instance = models.City.get(name=city['name'])
            if not city_instance:
                city_added_count += 1
                models.City.objects.get_or_create(
                    id=city['id'], name=city['name'], country=country_instance,
                    lon=str(city['coord']['lon']), lat=str(city['coord']['lat'])
                )
            c += 1
    print(f'All cities loaded: {city_added_count} cities and {country_added_count} countries added. '
          f'{total_count} cities checked.')


def load_forecasts_to_excel(forecasts):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Forecasts'
    columns = [
        ("A", "City ID", 10),
        ("B", "City name", 20),
        ("C", "Brief", 20),
        ("D", "JSON", 50),
    ]
    cols = 1
    row_count = 1
    for i in columns:
        ws.column_dimensions[i[0]].width = i[2]
        ws.cell(row=1, column=cols, value=i[1])
        cols += 1
    row_count += 1
    for forecast in forecasts:
        ws.cell(row=row_count, column=1, value=forecast.city.id)
        ws.cell(row=row_count, column=2, value=forecast.city.name)
        ws.cell(row=row_count, column=3, value=forecast.detailed_status)
        ws.cell(row=row_count, column=4, value=str(forecast.data))
        row_count += 1

    file_name = f"{get_now().strftime('%Y-%m-%d %H:%M:%S')}.xlsx"
    relative_path = f"files/{file_name}"
    path = os.path.join(os.getcwd(), relative_path)
    wb.save(path)
    return f'/{relative_path}'


def load_forecasts_from_api():
    owm = pyowm.OWM(settings.API_KEY)
    mgr = owm.weather_manager()
    cities = models.City.all()
    total_count = cities.count()
    c = 1
    for city in cities:
        observation = mgr.weather_at_id(city.id)
        w = observation.weather
        print(f'{c}/{total_count} Loading forecasts for {city.name}')
        models.Forecast.objects.get_or_create(
            city=city,
            detailed_status=w.detailed_status,
            time=timestamp_to_datetime(w.reference_time()),
            data=w.to_dict()
        )
        c += 1
    print('All data loaded')
